#==========================================
# Title:  Main file
# Author: Nguyen Vu Huan  -  NHY2HC 
# Date:   7 Jan 2019
#==========================================
import filecmp
import os
import logging
# import openpyxl module
import openpyxl
import sys
import re
import Gen_xml_define
import win32api
import subprocess
import Gui_GENTC
from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QMessageBox
from PyQt5.QtGui import QIcon
from PyQt5.QtCore import pyqtSlot
#catch error
check_error = None
#DEFINE COLUMN NAME
RequestResponseReg = 'RequestResponse\((.*)[ ]*,[ ]*(.*)[ ]*,[ ]*(.*)\)'
COL_TC_ID = "ID.?"
COL_component_name = "XXX Component MDC Component COM Tests MDC DCOM Tests LabT_MDC_ERRMGR MDC Error Manager Tests  LabT_MDC_DCOM MDC_ERRMGR Labor Test Lane Deviation Warning"
COL_test_description = "Test Description.?"
COL_test_Step = "Test.*Steps.?"
COL_test_response = "Test response.?"
COL_teststep_keywords = "Test.+?words"
COL_objectType = "ObjectType.?"
COL_teststatus = "TestStatus.?"
COL_project = "Project.?"
# todo --> modify input from gui
COL_release = ""
COL_test_result = ""
counter = 0
log_handle = logging.getLogger('ToolHuan')
glb_progress = None





class Data_store():
    def __init__(self):
        self.TC_ID_arr = list()
        self.Component_name_arr = list()
        self.test_description_arr = list()
        self.test_step_arr = list()
        self.project_arr = list()
        self.test_response_arr = list()
        self.teststep_key_arr = list()
        self.test_status_arr = list()
        self.object_type_arr = list()
        # self.test_release = None
        self.test_release_arr = list()
        self.test_result_arr = list()
        self.end_row = list()
        self.begin_ID = 0
        self.end_ID = 0


    def call_ID(TC_ID):
        # pass
        print(TC_ID)


# this function is get data from excel file
def Get_data(t_path_input="",t_release_hd="", t_test_result_hd="", t_range_TC="", t_release=""):
    catch_error = None
    check_error = None
    #Set work sheet to be gen
    # Get data from GUI
    # Give the location of the file
    path_input = t_path_input
    global COL_release
    COL_release = t_release_hd
    global COL_test_result
    COL_test_result = t_test_result_hd
    # define class object
    data_store = Data_store()
    # data_store.test_release = t_release
    Gen_xml_define.ProgressValue_obj_2.setValue(10)
    #     # calc range
    tt_range = [None, None]
    t_range = re.findall("\d+", t_range_TC)
    for i in range(len(t_range)):
        tt_range[i] = t_range[i]
    t_begin_ID = int(tt_range[0])
    if tt_range[1]:
        t_end_ID = int(tt_range[1])
    else:
        t_end_ID = 0
    # workbook object is created
    wb_obj = openpyxl.load_workbook(path_input)
    # sheet_obj = wb_obj['Sheet1']
    sheet_obj = wb_obj.active
    cell_obj = sheet_obj.cell(row=1, column=1)
    # log_handle.info(cell_obj.value)
    print(cell_obj.value)
    data_store.end_row = sheet_obj.max_row
    for i in range(1, sheet_obj.max_column+1):
        temp = str(sheet_obj.cell(row=1, column=i).value)
        if temp != None:
            if (re.search(COL_teststatus, temp, re.IGNORECASE)):
                # #if COL_teststatus.find(temp) != -1:
                for j in range(1, sheet_obj.max_row + 1):
                    data_store.test_status_arr.append(sheet_obj.cell(row=j, column=i).value)
            if (re.search(COL_TC_ID, temp, re.IGNORECASE)):
                for j in range(1, sheet_obj.max_row + 1):
                    data_store.TC_ID_arr.append(str(sheet_obj.cell(row=j, column=i).value))
            if (re.search(temp, COL_component_name, re.IGNORECASE)):
                #if COL_component_name.find(temp) != -1:
                for j in range(1, sheet_obj.max_row + 1):
                    data_store.Component_name_arr.append(sheet_obj.cell(row=j, column=i).value)
            if (re.search(COL_teststep_keywords,temp,  re.IGNORECASE)):
                # #if COL_teststep_keywords.find(temp) != -1:
                for j in range(1, sheet_obj.max_row + 1):
                    data_store.teststep_key_arr.append(sheet_obj.cell(row=j, column=i).value)
            if (re.search(COL_test_description, temp, re.IGNORECASE)):
                # if COL_test_description.find(temp) != -1:
                for j in range(1, sheet_obj.max_row + 1):
                    data_store.test_description_arr.append(sheet_obj.cell(row=j, column=i).value)
            if (re.search(COL_test_Step, temp, re.IGNORECASE)):
                # if COL_test_Step.find(temp) != -1:
                for j in range(1, sheet_obj.max_row + 1):
                    data_store.test_step_arr.append(sheet_obj.cell(row=j, column=i).value)
            if (re.search(COL_test_response, temp, re.IGNORECASE)):
                # if COL_test_response.find(temp) != -1:
                for j in range(1, sheet_obj.max_row + 1):
                    data_store.test_response_arr.append(sheet_obj.cell(row=j, column=i).value)
            if (re.search(COL_objectType, temp, re.IGNORECASE)):
                # if COL_objectType.find(temp) != -1:
                for j in range(1, sheet_obj.max_row + 1):
                    data_store.object_type_arr.append(sheet_obj.cell(row=j, column=i).value)
            if (re.search(COL_test_result, temp, re.IGNORECASE)):
                # if COL_test_result.find(temp) != -1:
                for j in range(1, sheet_obj.max_row + 1):
                    data_store.test_result_arr.append(sheet_obj.cell(row=j, column=i).value)
            if (re.search(COL_project, temp, re.IGNORECASE)):
                # if COL_project.find(temp) != -1:
                for j in range(1, sheet_obj.max_row + 1):
                    data_store.project_arr.append(sheet_obj.cell(row=j, column=i).value)
            if (re.search(COL_release, temp, re.IGNORECASE)):
                # if COL_release.find(temp) != -1:
                for j in range(1, sheet_obj.max_row + 1):
                    data_store.test_release_arr.append(sheet_obj.cell(row=j, column=i).value)
    Gen_xml_define.ProgressValue_obj_2.setValue(20)
    if len(data_store.test_status_arr) == 0:
        check_error = 1
        catch_error = "PLease check the test status name in first row (Should be TestStatus) "
    if len(data_store.TC_ID_arr) == 0:
        check_error = 1
        catch_error = "PLease check the TC_ID_arr in first row (Should be ID) "
    if len(data_store.Component_name_arr) == 0:
        for k in range(1,sheet_obj.max_row + 1):
            data_store.Component_name_arr.append(sheet_obj.cell(row=k, column=2).value)
        if len(data_store.Component_name_arr) == 0:
            check_error = 1
            catch_error = "PLease check the test Component_name in first row (Should be MDC Component or  Error Manager Tests or COM Tests)"
    if len(data_store.teststep_key_arr) == 0:
        check_error = 1
        catch_error = "PLease check the _teststep_keywords in first row (Should be Teststep keywords)"
    if len(data_store.test_description_arr) == 0:
        check_error = 1
        catch_error = "PLease check the test_description in first row (should be Test Description)"
    if len(data_store.test_step_arr) == 0:
        check_error = 1
        catch_error = "PLease check the test_step in first row (should be TestSteps)"
    if len(data_store.test_response_arr) == 0:
        check_error = 1
        catch_error = "PLease check the test_response in first row (should be Test response)"
    if len(data_store.object_type_arr) == 0:
        check_error = 1
        catch_error = "PLease check the object_type in first row (should be ObjectType)"
    if len(data_store.project_arr) == 0:
        check_error = 1
        catch_error = "PLease check the project name in first row ,have to add Project column when export from DOOR  (should be Project)"
    if (check_error != None):
        win32api.MessageBox(0, catch_error, 'ERROR')

    data_store.begin_ID = t_begin_ID-1
    data_store.end_ID = t_end_ID
    return data_store


def Write_data_to_DLX_file(t_path_input, t_path_output, t_release_hd, t_test_result_hd, t_range_TC, t_test_status, t_Testmethods, t_TestEnvironment,t_TestreviewStatus, t_TestFrequency, t_SafetyRequirement, t_TestSequence):
    catch_error = None
    check_error = None
    if os.path.exists(t_path_output) and os.path.isfile(t_path_input):
    # Get data from excel file
        data_w = Get_data(t_path_input, t_release_hd, t_test_result_hd, t_range_TC)
        path_output = t_path_output
        # ============================================================
        t_Testmethods = re.findall('[ \w]+', t_Testmethods)
        #remove unneed space
        for i in range (len(t_Testmethods)):
            t_Testmethods[i] = t_Testmethods[i].lstrip()

        t_test_status = re.findall('[ \w]+', t_test_status)
        #remove unneed space
        for i in range (len(t_test_status)):
            t_test_status[i] = t_test_status[i].lstrip()
        #=============================================================
        # input data
        if len(data_w.test_release_arr) == 0:
            check_error = 1
            catch_error = "PLease check the release header in first row and the input in tool"
        if len(data_w.test_result_arr) == 0:
            check_error = 1
            catch_error = "PLease check the test result header in first row and the input in tool"
        if (check_error != None):
            win32api.MessageBox(0, catch_error, 'ERROR')
        Gen_xml_define.ProgressValue_obj_2.setValue(30)
        RC_release_maximum = 0
        if data_w.end_ID == 0:
            number_tc = (len(data_w.TC_ID_arr) - data_w.begin_ID)
            data_w.end_ID = len(data_w.TC_ID_arr)
        else:
            number_tc = data_w.end_ID - data_w.begin_ID
        absolute_number_arr = []
        object_heading_arr = []
        while len(absolute_number_arr) < data_w.begin_ID:
            absolute_number_arr.append("none")
            object_heading_arr.append("none")
        # using regex to get only number
        if data_w.end_ID > len(data_w.TC_ID_arr):
            catch_error = "Please check, Max Range of TC is "+ str(len(data_w.TC_ID_arr))
            win32api.MessageBox(0, catch_error, 'ERROR')
        else:
            for i in range(data_w.begin_ID, data_w.end_ID):
                temp = re.search("\d+$", (data_w.TC_ID_arr[i]))
                temp_2 = None
                if temp:
                    absolute_number_arr.append(temp.group())
                else:
                    absolute_number_arr.append(None)
                if (data_w.Component_name_arr[i]):
                    temp_2 = re.findall("[a-z].*", (data_w.Component_name_arr[i]), re.IGNORECASE)
                    temp_2 = '\n'.join(temp_2)
                if temp_2:
                    object_heading_arr.append(temp_2)
                else:
                    object_heading_arr.append(None)
            path_output = path_output + r'\updoor.txt'
            with open(path_output, 'w',encoding="utf-8") as file:
                # out value
                temp_arr = list()
                Gen_xml_define.ProgressValue_obj_2.setValue(40)
                # file.write("\n".join(abc))

                file.write("print \"Start...\\n\"" + "\n")
                file.write("int NumberTestCase = " + str(number_tc)+ "\n")
                file.write(r"//Array of Objects with its " + "\n")
                file.write("Array ArrObjectInfor = create(NumberTestCase," + str(number_tc*20) + ")" + "\n")
                file.write("const int Absolute_number = 0" + "\n")
                file.write("const int Object_Heading = 1" + "\n")
                file.write("const int Test_Description = 2" + "\n")
                file.write("const int Test_Step = 3" + "\n")
                file.write("const int Test_Response = 4" + "\n")
                file.write("const int Teststep_keywords = 5" + "\n")
                file.write("const int ObjectType = 6" + "\n")
                file.write("const int TestStatus = 7" + "\n")
                # file.write("const int Release = 8" + "\n")
                file.write("const int TestResult = 8" + "\n")
                file.write("const int TestEnvironment = 9" + "\n")
                file.write("const int TestreviewStatus = 10" + "\n")
                file.write("const int Test_Frequency = 11" + "\n")
                file.write("const int Safety_Requirement = 12" + "\n")
                file.write("const int Test_Sequence = 13" + "\n")
                file.write("const int tProject = 14" + "\n")
                counter = 0
                for k in range(len(t_Testmethods)):
                    if t_Testmethods[k]:
                        counter = k+1+14
                        file.write("const int Test_methods_"+ str(k+1)+" = "+ str(counter)  + "\n")
                file.write("const int Release = " + str(15+ len(t_Testmethods)) + "\n")
                file.write(r"// row: mapped with testcase ID" + "\n")
                file.write(r"//Colunm: " + "\n")
                file.write(r"//	0: Contains of Absolute number" + "\n")
                file.write(r"//	1: Contains of Object Heading" + "\n")
                file.write(r"//	2: Contains of Test Description" + "\n")
                file.write(r"//	3: Contains of TestSteps" + "\n")
                file.write(r"//	4: Contains of Test Response" + "\n")
                file.write(r"//  5: Contains of Teststep keywords" + "\n")
                file.write(r"//  6: Contains of ObjectType" + "\n")
                file.write(r"//  7: Contains of TestStatus" + "\n")
                file.write(r"//  8: Contains of RC_DFSK_F517_R1.0" + "\n")
                file.write(r"//  9: Contains of TestResult_DFSK_F517_R1.0" + "\n")
                j = 0
                for i in range(data_w.begin_ID, data_w.end_ID):
                    if absolute_number_arr[i]:
                        file.write(r"   put(ArrObjectInfor," + absolute_number_arr[i] + "," + str(j) + ",Absolute_number)\n")
                    if object_heading_arr[i]:
                        file.write(r"   put(ArrObjectInfor," + "\"" + object_heading_arr[i] + "\"" + "," + str(j) + ",Object_Heading)\n")
                    test_status_check ="None"
                    # for l in range (len(t_test_status)):
                    #     if (data_w.test_status_arr[i] == t_test_status[l]):
                    #         test_status_check  = 1;
                    #
                    # if(test_status_check == 1):

                    if data_w.test_description_arr[i]:
                        test_Ss = str(data_w.test_description_arr[i])
                        rexReplace = re.sub(r"\"", r"\"", (re.sub(r"\\", r"\\\\", test_Ss)))
                        file.write(r"   put(ArrObjectInfor," + "\"" + rexReplace + "\"" + "," + str(j) + ",Test_Description)\n")
                    if data_w.test_step_arr[i]:
                        rexReplace = re.sub(r"\"", r"\"", (re.sub(r"\\", r"\\\\", data_w.test_step_arr[i])))
                        file.write(r"   put(ArrObjectInfor," + "\"" + data_w.test_step_arr[i] + "\"" + "," + str(j) + ",Test_Step)\n")
                    if data_w.test_response_arr[i]:
                        file.write(r"   put(ArrObjectInfor," + "\"" + data_w.test_response_arr[i] + "\"" + "," + str(j) + ",Test_Response)\n")
                    if data_w.teststep_key_arr[i]:
                        rexReplace = re.sub(r"\"", r"\"",(re.sub(r"\\",r"\\\\", data_w.teststep_key_arr[i])))
                        file.write(r"   put(ArrObjectInfor," + "\"" + rexReplace + "\"" + "," + str(j) + ",Teststep_keywords)\n")
                    if data_w.object_type_arr[i]:
                        rexReplace = re.sub(r"\"", r"\"", (re.sub(r"\\", r"\\\\", data_w.object_type_arr[i])))
                        file.write(r"   put(ArrObjectInfor," + "\"" + data_w.object_type_arr[i] + "\"" + "," + str(j) + ",ObjectType)\n")
                    if data_w.test_status_arr[i]:
                        rexReplace = re.sub(r"\"", r"\"", (re.sub(r"\\", r"\\\\", data_w.test_status_arr[i])))
                        file.write(r"   put(ArrObjectInfor," + "\"" + data_w.test_status_arr[i] + "\"" + "," + str(j) + ",TestStatus)\n")
                    if data_w.test_release_arr[i]:
                        t_split = str(data_w.test_release_arr[i]).split("\n")
                        RC_arr = list()
                        for m in range(len(t_split)):
                            # t_RC_Split = t_split[j]
                            if (t_split[m]):
                                RC_arr.append(re.sub(' ', '',t_split[m] ))
                        if (RC_release_maximum< len(RC_arr)):
                            RC_release_maximum = len(RC_arr)
                        for m in range (len(RC_arr)):
                            # if (re.search(r'RC\d\d',RC_arr[m])):
                            file.write(r"   put(ArrObjectInfor," + "\"" + RC_arr[m] + "\"" + "," + str(j) + ",Release +" + str(m)+ ")\n")
                    if (data_w.test_result_arr[i] == "failed"):
                        file.write(r"   put(ArrObjectInfor," + "\"" + "failed" + "\"" + "," + str(j) + ",TestResult)\n")
                    elif (data_w.test_result_arr[i] == "passed"):
                        file.write(r"   put(ArrObjectInfor," + "\"" + "passed" + "\"" + "," + str(j) + ",TestResult)\n")
                    for k in range (len(t_Testmethods)):
                        if t_Testmethods[k]:
                            file.write(r"   put(ArrObjectInfor," + "\"" + t_Testmethods[k] + "\"" + "," + str(j) + ",Test_methods_"+ str(k+1)+")\n")
                    if t_TestEnvironment:
                        if (data_w.test_step_arr[i]) != None:
                            if (re.findall('Bus off', data_w.test_step_arr[i])):
                                file.write(r"   put(ArrObjectInfor," + "\"" + "CANStress" + "\"" + "," + str(j) + ",TestEnvironment)\n")
                            else:
                                file.write(r"   put(ArrObjectInfor," + "\"" + t_TestEnvironment + "\"" + "," + str(j) + ",TestEnvironment)\n")
                    if t_TestreviewStatus:
                        if(data_w.test_status_arr[i] == "specified" ):
                            file.write(r"   put(ArrObjectInfor," + "\"" + "specification_reviewed" + "\"" + "," + str(j) + ",TestreviewStatus)\n")
                        elif(data_w.test_status_arr[i] == "implemented" ):
                            file.write(r"   put(ArrObjectInfor," + "\"" + "implementation_reviewed" + "\"" + "," + str(j) + ",TestreviewStatus)\n")
                    if t_TestFrequency:
                        if (data_w.object_type_arr[i] == "Automated Testcase"):
                            file.write(r"   put(ArrObjectInfor," + "\"" + t_TestFrequency + "\"" + "," + str(j) + ",Test_Frequency)\n")
                    if t_SafetyRequirement:
                        file.write(r"   put(ArrObjectInfor," + "\"" + t_SafetyRequirement + "\"" + "," + str(j) + ",Safety_Requirement)\n")
                    if t_TestSequence:
                        file.write(r"   put(ArrObjectInfor," + "\"" + t_TestSequence + "\"" + "," + str(j) + ",Test_Sequence)\n")
                    if data_w.project_arr[i]:
                        file.write(r"   put(ArrObjectInfor," + "\"" + data_w.project_arr[i] + "\"" + "," + str(j) + ",tProject)\n")

                    file.write("\n")
                    j = j+1

                file.write("Module ModTemp" + "\n")
                file.write("string StrTemp" + "\n")
                file.write("int temp1																								 " + "\n")
                file.write("		Object currOject                                                                                 " + "\n")
                file.write("		Object ObjTemp                                                                                   " + "\n")
                file.write("		bool CheckOValid                                                                                 " + "\n")
                file.write("		int CurrTestCaseAbsID                                                                            " + "\n")
                file.write("		                                                                                                 " + "\n")
                file.write("for (loop = 0; loop<NumberTestCase; loop++)                                                             " + "\n")
                file.write("		{                                                                                                " + "\n")
                file.write("			CurrTestCaseAbsID = (int get(ArrObjectInfor,loop,Absolute_number))                           " + "\n")
                file.write("			print CurrTestCaseAbsID \"\\n\"                                                                 " + "\n")
                file.write("			CheckOValid = false                                                                          " + "\n")
                file.write("			for ObjTemp in current Module do                                                             " + "\n")
                file.write("			{                                                                                            " + "\n")
                file.write("				temp1 = ObjTemp.\"Absolute Number\"                                                        " + "\n")
                file.write("				if (temp1==CurrTestCaseAbsID) CheckOValid = true                                         " + "\n")
                file.write("			}                                                                                            " + "\n")
                file.write("			if(CheckOValid == true)                                                                      " + "\n")
                file.write("			{                                                                                            " + "\n")
                file.write("			currOject = object(CurrTestCaseAbsID)                                                        " + "\n")
                file.write("			print CurrTestCaseAbsID \"\\n\"                                                                 " + "\n")
                file.write("			// Added lines to update requirement with absolutely same as testcase                        " + "\n")
                file.write("	                                                                                                     " + "\n")
                file.write("			}                                                                                            " + "\n")
                file.write("			else                                                                                         " + "\n")
                file.write("			{                                                                                            " + "\n")
                file.write("				CurrTestCaseAbsID = (int get(ArrObjectInfor,loop-1,Absolute_number))                     " + "\n")
                file.write("				currOject = create after object(CurrTestCaseAbsID)                                       " + "\n")
                file.write("				print currOject.\"Absolute number\" \"\\n\"                                                   " + "\n")
                file.write("			}                                                                                            " + "\n")
                file.write("		                                                                                                 " + "\n")
                file.write("			//update all attribute to Current Object if status is implemented or specified                                                  " + "\n")
                file.write("			                                                                                             " + "\n")
                file.write("			string temp_check                                                                                             " + "\n")
                file.write("			temp_check = (string get(ArrObjectInfor,loop,TestStatus))                                                                                          " + "\n")
                # file.write("			if((temp_check != \"implemented\") && (temp_check != \"specified\"))                                                                                            " + "\n")
                # file.write("			{                                                                                             " + "\n")
                # file.write("			//donothing                                                                                             " + "\n")
                # file.write("			}                                                                                            " + "\n")
                # file.write("			else                                                                                            " + "\n")
                # file.write("			{                                                                                            " + "\n")
                file.write("			currOject.\"Object Heading\"   			 = (string get(ArrObjectInfor,loop,Object_Heading))   " + "\n")
                file.write("			currOject.\"Test Description\"  			= (string get(ArrObjectInfor,loop,Test_Description)) " + "\n")
                file.write("			currOject.\"TestSteps\" 					= (string get(ArrObjectInfor,loop,Test_Step))        " + "\n")
                file.write("			currOject.\"Test Response\" 				= (string get(ArrObjectInfor,loop,Test_Response))	 " + "\n")
                file.write("			currOject.\"Teststep keywords\" 			= (string get(ArrObjectInfor,loop,Teststep_keywords))" + "\n")
                file.write("			currOject.\"ObjectType\" 					= (string get(ArrObjectInfor,loop,ObjectType))       " + "\n")
                for k in range(len(t_Testmethods)):
                    if t_Testmethods[k]:
                        file.write("			currOject.\"Test methods\" 		        += (string get(ArrObjectInfor,loop,Test_methods_"+str(k+1)+"))       " + "\n")
                file.write("			currOject.\"TestEnvironment\" 			= (string get(ArrObjectInfor,loop,TestEnvironment))       " + "\n")
                file.write("			currOject.\"TestreviewStatus\" 			= (string get(ArrObjectInfor,loop,TestreviewStatus))       " + "\n")
                file.write("			currOject.\"Test Frequency\" 				= (string get(ArrObjectInfor,loop,Test_Frequency))       " + "\n")
                file.write("			currOject.\"Safety Requirement\" 			= (string get(ArrObjectInfor,loop,Safety_Requirement))       " + "\n")
                file.write("			currOject.\"Test Sequence\" 				= (string get(ArrObjectInfor,loop,Test_Sequence))       " + "\n")
                file.write("			currOject.\"ObjectType\" 					= (string get(ArrObjectInfor,loop,ObjectType))       " + "\n")

                file.write("			currOject.\"TestStatus\" 					= (string get(ArrObjectInfor,loop,TestStatus))       " + "\n")
                for k in range(RC_release_maximum):
                    file.write("			if ((string get(ArrObjectInfor,loop,Release +" + str(k)+"))""!= null)                                                                     " + "\n")
                    file.write("			{                                                                                            " + "\n")
                    file.write(" 			currOject.\"" + COL_release + "\" 			    += (string get(ArrObjectInfor,loop,Release +" + str(k)+ "))          " + "\n")
                    file.write("			}                                                                                            " + "\n")
                file.write("			if ((string get(ArrObjectInfor,loop,TestResult ))""!= null)                                                                     " + "\n")
                file.write("			{                                                                                            " + "\n")
                file.write(" 			currOject.\"" + COL_test_result     + "\"      = (string get(ArrObjectInfor,loop,TestResult))       " + "\n")
                file.write("			}                                                                                            " + "\n")

                file.write("			if ((string get(ArrObjectInfor,loop,tProject ))""!= null)                                                                     " + "\n")
                file.write("			{                                                                                            " + "\n")
                file.write("			currOject.\"Project\" 		            += (string get(ArrObjectInfor,loop,tProject))       " + "\n")
                file.write("			}                                                                                            " + "\n")

                # file.write("			}                                                                                            " + "\n")
                file.write("			                                                                                             " + "\n")
                file.write("			//link current object to its requirement                                                     " + "\n")
                file.write("			int loop1 = 0 // start at location stores requirement                                        " + "\n")
                file.write("			int ReqIDCurr = 0                                                                            " + "\n")
                file.write("			Object ObjTemp1                                                                              " + "\n")
                file.write("			bool CheckOReqValid                                                                          " + "\n")
                file.write("	 }" + "\n")
                Gen_xml_define.ProgressValue_obj_2.setValue(70)

    else:
        if os.path.exists(t_path_output):
            catch_error = "please check input Testcase path"
            check_error = 1
        elif os.path.isfile(t_path_input):
            catch_error = "please check Output  path"
            check_error = 1
        else:
            catch_error = "please check INPUT and Output path"
            check_error = 1
        Gen_xml_define.ProgressValue_obj_2.setValue(50)
    if (check_error != None):
        win32api.MessageBox(0, catch_error, 'ERROR')
        Gen_xml_define.ProgressValue_obj_2.setValue(90)
    path_output = t_path_output + r'\updoor.txt'
    with open(path_output, 'r') as file:
        content = file.read()
        content = re.sub("aka\d+_", "", content)
        content = re.sub("_x000D_", "", content)
    with open(path_output, "w") as f:
        f.write(content)
    cmd_command = r'"C:\Program Files\Notepad++\notepad++.exe" {}'.format(path_output)
    notepad_path_check = r'C:\Program Files\Notepad++\notepad++.exe'
    Gen_xml_define.ProgressValue_obj_2.setValue(100)
    if (os.path.exists(notepad_path_check) == True):
        # print(cmd_command)
        subprocess.Popen(cmd_command)

    else:
        cmd_command = r'"C:\Program Files (x86)\Notepad++\notepad++.exe" {}'.format(path_output)
        notepad_path_check = r'C:\Program Files (x86)\Notepad++\notepad++.exe'
        if (os.path.exists(notepad_path_check) == True):
            # print(cmd_command)
            subprocess.Popen(cmd_command)

        else:
            catch_error = r'Please check the notepad folder (should be "C:\Program Files\Notepad++\notepad++.exe")'
            win32api.MessageBox(0, catch_error, 'ERROR')



def Generate_test_case_to_xlm(t_path_input, t_path_output, t_range_TC, t_ECAN_main_node, t_SFCAN_main_node, t_measure_type, t_test_module, t_execute_option,t_Generate_option_select):
    catch_error = None
    check_error = None
    if os.path.exists(t_path_output) and os.path.isfile(t_path_input):
        data_w = Get_data(t_path_input=t_path_input, t_range_TC=t_range_TC, )
        path_output = t_path_output
        object_heading_arr = list()
        while len(object_heading_arr) < 1:
            object_heading_arr.append("none")
        for i in range(1, data_w.end_row):
            if data_w.Component_name_arr[i]:
                object_heading_arr.append(Gen_xml_define.object_heading())
                temp_1 = re.findall('^\s*([\d.]+? )',(data_w.Component_name_arr[i]), re.IGNORECASE)
                if (temp_1):
                    temp_1 = re.findall('[0-9]+', temp_1[0], re.IGNORECASE)
                    check_error = None
                    if (data_w.object_type_arr[i]!="Test group" and  data_w.teststep_key_arr[i]==None and data_w.test_step_arr[i] == None and data_w.test_response_arr[i]== None):
                        if data_w.object_type_arr[i] == None or data_w.object_type_arr[i] == "Automated Testcase" or data_w.object_type_arr[i] =="Manual Testcase":
                            check_error = 1
                            catch_error = "Object type  may be empty or wrong, Please check row: " + str(i + 1)
                            break
                        else:
                            data_w.object_type_arr[i]="Test group"

                else:
                    check_error = 1
                    catch_error = "Component_name_arr may be empty or wrong, Please check row: " + str(i+1)
                    break
                temp_2 = re.findall("[a-z].*", (data_w.Component_name_arr[i]), re.IGNORECASE)

                if temp_2:
                    object_heading_arr[i].suffix.append(temp_2[0])
                else:
                    object_heading_arr[i].suffix.append(None)
                if temp_1:
                    object_heading_arr[i].prefix.append(temp_1)
                else:
                    object_heading_arr[i].prefix.append(None)
            else:
                object_heading_arr.append(None)
        # step_by_step_keyw = Gen_xml_define.Step_by_step_keyw()
        # step_by_test_step = Gen_xml_define.Step_by_test_step()
        # step_by_test_response = Gen_xml_define.Step_by_test_response()
        if (check_error == None):
            step_by_step_keyw_arr = list()
            step_by_test_step_arr = list()
            step_by_test_response_arr = list()
            while len(step_by_step_keyw_arr) < 1:
                step_by_step_keyw_arr.append("None")
                step_by_test_step_arr.append("None")
                step_by_test_response_arr.append("None")
            temp = None
            temp_2 = None
            # input data
            for i in range(1, data_w.end_row):
                step_by_step_keyw_arr.append(Gen_xml_define.Step_by_step_keyw())
                step_by_test_step_arr.append(Gen_xml_define.Step_by_test_step())
                step_by_test_response_arr.append(Gen_xml_define.Step_by_test_response())
                if (data_w.teststep_key_arr[i] != None):
                    data_w.teststep_key_arr[i] = re.sub('\r\n//','',data_w.teststep_key_arr[i])
                    # data_w.teststep_key_arr[i] = '\n'.join(data_w.teststep_key_arr[i])
                t_split =str(data_w.teststep_key_arr[i]).split("\n")
                for j in range(len(t_split)):
                    t_step_check = re.findall('^\d+\)', t_split[j])
                    if (t_step_check):
                        step_by_step_keyw_arr[i].step_contain.append(re.sub('','',re.sub('^\d+\)','',t_split[j])))
                        step_by_step_keyw_arr[i].step_check.append(re.findall('^\d',t_step_check[0]))
                        step_by_step_keyw_arr[i].step_count = step_by_step_keyw_arr[i].step_count+1

                t_split = str(data_w.test_step_arr[i]).split("\n")
                for j in range(len(t_split)):
                    t_step_check = re.findall('^\d+\)', t_split[j])
                    if (t_step_check):
                        step_by_test_step_arr[i].step_contain.append(re.sub('^ ', '', re.sub('^\d+\)', '', t_split[j])))
                        step_by_test_step_arr[i].step_check.append(re.findall('^\d', t_step_check[0]))
                        step_by_test_step_arr[i].step_count = step_by_test_step_arr[i].step_count + 1

                t_split = str(data_w.test_response_arr[i]).split("\n")
                for j in range(len(t_split)):
                    t_step_check = re.findall('^\d+\)', t_split[j])
                    if (t_step_check):
                        step_by_test_response_arr[i].step_contain.append(re.sub('^ ', '', re.sub('^\d+\)', '', t_split[j])))
                        step_by_test_response_arr[i].step_check.append(re.findall('^\d', t_step_check[0]))
                        step_by_test_response_arr[i].step_count = step_by_test_response_arr[i].step_count + 1

                Gen_xml_define.ProgressValue_obj.setValue(10)
            # Check step correct or not
                if(data_w.object_type_arr[i] == "Automated Testcase")and ((data_w.test_status_arr[i] == "implemented")or (data_w.test_status_arr[i] == "specified")):
                    if (step_by_step_keyw_arr[i].step_contain):
                        for j in range(step_by_step_keyw_arr[i].step_count):
                            if (step_by_step_keyw_arr[i].step_count == step_by_test_step_arr[i].step_count == step_by_test_response_arr[i].step_count):
                                if (step_by_step_keyw_arr[i].step_check[j] == step_by_test_step_arr[i].step_check[j] ==step_by_test_response_arr[i].step_check [j]):
                                    check_error = None
                                    temp_2 = j+1
                                else:
                                    check_error = 1
                                    catch_error = "Generate may be error, Please check step " + str(temp_2) + ",Test row: "+ str(i+1)
                                    break
                            else:
                                check_error = 1
                                catch_error = "Generate may be error, Please count of step " + str(step_by_step_keyw_arr[i].step_count) + ",Test row: "+ str(i+1)
                                break
                    else:
                        check_error = 1
                        catch_error = "Generate may be error, Please check Test step keyword column, row: " + str(i + 1)

                else:
                    check_error = None
            # show error
                if (check_error != None):
                    win32api.MessageBox(0, catch_error, 'ERROR')

            Gen_xml_define.ProgressValue_obj.setValue(20)
            Gen_xml_define.write_xml(data_w, step_by_step_keyw_arr, step_by_test_step_arr, step_by_test_response_arr, path_output, object_heading_arr, t_ECAN_main_node, t_SFCAN_main_node, t_measure_type, t_test_module,t_execute_option,t_Generate_option_select)


    else:
        if os.path.exists(t_path_output):
            catch_error = "please check input Testcase path"
            check_error = 1
        elif os.path.isfile(t_path_input):
            catch_error = "please check Output  path"
            check_error = 1
        else:
            catch_error = "please check INPUT and Output path"
            check_error = 1
    if (check_error != None):
        win32api.MessageBox(0, catch_error, 'ERROR')


if __name__ == '__main__':
    # *********************************************************
    # INPUT

    t_release_hd = "RC_DFSK"
    t_test_result_hd = "Test_Result"
    t_range_TC = 3
    t_release = "RC01"
    t_path_output = r"C:\Users\UAN1HC\Desktop\aka.txt"
    t_path_input = r"C:\Users\UAN1HC\Desktop\TC.xlsx"
    # *********************************************************
    # Get_data(1,2,3)
    Write_data_to_DLX_file(t_path_output, t_path_input, t_release_hd, t_test_result_hd, t_range_TC,t_release)
    Generate_test_case_to_xlm(t_path_input, t_path_output, t_range_TC)

    # checkupdate()

